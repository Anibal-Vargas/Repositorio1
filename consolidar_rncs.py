#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
consolidar_rncs.py — consolida em uma planilha Excel as não conformidades da
categoria "Instalações elétricas em más condições" encontradas nos relatórios
de inspeção (.pdf e .docx) de uma pasta, eliminando duplicatas entre relatórios.

Fluxo:
 1. Lista os .pdf/.docx da pasta e pede confirmação antes de prosseguir;
 2. Em cada relatório, localiza a seção cujo título é igual ou semelhante a
    "Instalações elétricas em más condições" (variações aceitas — ver
    VARIANTES_TITULO); relatórios sem a seção são pulados e registrados no log;
 3. Extrai da seção as não conformidades e seus subitens, capturando fielmente
    descrição, item de norma (NR-10 / ABNT) e solução proposta — campos ausentes
    recebem "[NÃO CONSTA NO RELATÓRIO]"; trechos ilegíveis recebem "[REVISAR]";
 4. Consolida duplicatas entre relatórios (mesma NC em vários relatórios vira
    uma linha, mantendo a redação mais completa, a frequência e as origens);
 5. Mostra as 5 primeiras linhas para validação e, após confirmação, grava
    nao_conformidades_consolidadas.xlsx e log_processamento.txt na pasta.

Regras de segurança:
 - os relatórios originais são abertos SOMENTE para leitura, nunca alterados;
 - PDFs escaneados (sem texto selecionável) NÃO são adivinhados: entram no log
   como "requer OCR" e o script pergunta se deve continuar sem eles;
 - nada é sobrescrito sem confirmação explícita.
"""

import argparse
import difflib
import os
import re
import sys
import unicodedata
from dataclasses import dataclass, field

PASTA_PADRAO = r"C:\Temp\RNCs\Compactados"
NOME_PLANILHA = "nao_conformidades_consolidadas.xlsx"
NOME_LOG = "log_processamento.txt"

NAO_CONSTA = "[NÃO CONSTA NO RELATÓRIO]"
REVISAR = "[REVISAR]"

# Título da categoria e variações aceitas (comparadas sem acentos/maiúsculas)
VARIANTES_TITULO = [
    "instalacoes eletricas em mas condicoes",
    "instalacoes em mas condicoes de conservacao",
    "instalacoes eletricas em mau estado",
    "instalacoes eletricas em mas condicoes de conservacao",
    "instalacoes eletricas em ma conservacao",
    "instalacoes eletricas em mau estado de conservacao",
]
# Palavras-chave para variações não previstas: o título precisa conter
# "instalac" E uma das expressões de condição abaixo
PALAVRAS_CONDICAO = ("mas condicoes", "mau estado", "ma conservacao",
                     "mas condicao", "condicoes precarias")
LIMIAR_TITULO = 0.80      # similaridade mínima com uma variante conhecida
LIMIAR_DUPLICATA = 0.85   # similaridade mínima para consolidar duas NCs

# Rótulos (sem acento, minúsculos) que introduzem os campos dentro de uma NC
ROTULOS_NORMA = ("item da norma tecnica", "item da norma", "embasamento tecnico",
                 "embasamento normativo", "embasamento", "referencia normativa",
                 "base normativa", "fundamentacao", "normas", "norma")
ROTULOS_SOLUCAO = ("solucao proposta", "solucao", "acao corretiva",
                   "acao proposta", "acoes corretivas", "medida corretiva",
                   "medidas corretivas", "correcao proposta", "recomendacao",
                   "providencia")

# Referências normativas reconhecidas quando não há rótulo explícito
RE_NORMA = re.compile(
    r"(?:ABNT\s*)?NBR\s*-?\s*\d{3,6}(?:[:/]\d{4})?(?:\s*,?\s*itens?\s*[\d.]+[\d.,\se]*)?"
    r"|NR\s*-?\s*10\b(?:\s*,?\s*itens?\s*[\d.]+[\d.,\se]*)?",
    re.IGNORECASE)

# Linhas de legenda de foto/figura — ignoradas na extração
RE_FOTO = re.compile(r"^\s*(foto|figura|imagem|registro fotografico)s?\b",
                     re.IGNORECASE)

# Início de item numerado: "4.2.1", "4.2.1.", "4.2.1)" etc.
RE_ITEM_NUMERADO = re.compile(r"^\s*(\d+(?:\.\d+)+)[.)]?\s+(.*)")
# Marcadores de lista (nível principal e subitem)
MARCADORES_NIVEL0 = ("•", "-", "–", "*", "▪")
MARCADORES_NIVEL1 = ("◦", "‣", "·")
RE_LETRA = re.compile(r"^\s*[a-z]\)\s+(.*)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Utilidades de texto
# ---------------------------------------------------------------------------

def normalizar(texto):
    """minúsculas, sem acentos, espaços colapsados — para comparações."""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().lower()


def sem_numeracao(texto):
    """Remove numeração/marcador inicial de um título ("4.2 Título" -> "Título")."""
    m = RE_ITEM_NUMERADO.match(texto)
    if m:
        return m.group(2)
    return texto.lstrip("".join(MARCADORES_NIVEL0) + "".join(MARCADORES_NIVEL1) + " \t")


def texto_ilegivel(texto):
    """Heurística de trecho ilegível: caracteres de substituição ou quase nada legível."""
    if "�" in texto:
        return True
    legiveis = sum(c.isalnum() for c in texto)
    return len(texto) >= 8 and legiveis < len(texto) * 0.3


# ---------------------------------------------------------------------------
# Modelo
# ---------------------------------------------------------------------------

@dataclass
class Bloco:
    """Um trecho do documento: parágrafo (texto) ou tabela (linhas de células)."""
    texto: str = ""
    nivel_titulo: int = 0          # >0 quando o estilo/numeração indica título
    tabela: list = None            # lista de linhas; cada linha = lista de células


@dataclass
class NC:
    descricao: str
    norma: str
    solucao: str
    nivel: int                     # 0 = principal, 1 = subitem
    arquivo: str
    revisar: bool = False
    subitens: list = field(default_factory=list)


@dataclass
class Grupo:
    """Uma linha consolidada: mesma NC vista em um ou mais relatórios."""
    descricao: str
    norma: str
    solucao: str
    arquivos: list
    revisar: bool = False
    subgrupos: list = field(default_factory=list)

    def absorver(self, nc):
        if len(nc.descricao) > len(self.descricao):
            self.descricao = nc.descricao
        for atual, novo in (("norma", nc.norma), ("solucao", nc.solucao)):
            valor = getattr(self, atual)
            if novo != NAO_CONSTA and (valor == NAO_CONSTA or len(novo) > len(valor)):
                setattr(self, atual, novo)
        if nc.arquivo not in self.arquivos:
            self.arquivos.append(nc.arquivo)
        self.revisar = self.revisar or nc.revisar


# ---------------------------------------------------------------------------
# Extração de blocos (.docx e .pdf)
# ---------------------------------------------------------------------------

def blocos_docx(caminho):
    import docx
    documento = docx.Document(caminho)
    blocos = []
    for item in documento.iter_inner_content():
        if hasattr(item, "rows"):                       # Tabela
            linhas = [[c.text.strip() for c in linha.cells] for linha in item.rows]
            blocos.append(Bloco(tabela=linhas))
        else:                                           # Parágrafo
            texto = item.text.strip()
            if not texto:
                continue
            nivel = 0
            estilo = (item.style.name or "") if item.style else ""
            m = re.match(r"(?:heading|t[ií]tulo)\s*(\d+)", estilo, re.IGNORECASE)
            if m:
                nivel = int(m.group(1))
            blocos.append(Bloco(texto=texto, nivel_titulo=nivel))
    return blocos


def blocos_pdf(caminho):
    """Retorna (blocos, total_caracteres). Poucos caracteres => PDF escaneado."""
    from pypdf import PdfReader
    leitor = PdfReader(caminho)
    blocos, total = [], 0
    for pagina in leitor.pages:
        texto = pagina.extract_text() or ""
        total += len(texto.strip())
        for linha in texto.splitlines():
            linha = linha.strip()
            if linha:
                blocos.append(Bloco(texto=linha))
    return blocos, total


# ---------------------------------------------------------------------------
# Localização da seção da categoria
# ---------------------------------------------------------------------------

def eh_titulo_categoria(texto):
    limpo = normalizar(sem_numeracao(texto))
    if not limpo or len(limpo) > 120:
        return False
    if any(v in limpo for v in VARIANTES_TITULO):
        return True
    if "instalac" in limpo and any(p in limpo for p in PALAVRAS_CONDICAO):
        return True
    return max(difflib.SequenceMatcher(None, limpo, v).ratio()
               for v in VARIANTES_TITULO) >= LIMIAR_TITULO


def parece_titulo(bloco):
    """(é título?, profundidade) — por estilo, numeração ou caixa alta."""
    if bloco.tabela is not None:
        return False, 0
    if bloco.nivel_titulo:
        return True, bloco.nivel_titulo
    m = RE_ITEM_NUMERADO.match(bloco.texto)
    if m and len(bloco.texto) <= 120:
        return True, m.group(1).count(".") + 1
    t = bloco.texto
    if (len(t) <= 90 and t == t.upper() and sum(c.isalpha() for c in t) >= 5
            and not RE_FOTO.match(t)):
        return True, 1
    return False, 0


def localizar_secao(blocos, notas):
    """Devolve os blocos internos da seção da categoria, ou None se não houver."""
    inicio = profundidade = None
    for i, bloco in enumerate(blocos):
        if bloco.tabela is None and eh_titulo_categoria(bloco.texto):
            _, prof = parece_titulo(bloco)
            inicio, profundidade = i, (prof or 1)
            break
    if inicio is None:
        return None
    fim = len(blocos)
    for j in range(inicio + 1, len(blocos)):
        eh_tit, prof = parece_titulo(blocos[j])
        if eh_tit and prof <= profundidade and not eh_titulo_categoria(blocos[j].texto):
            fim = j
            break
    if fim == len(blocos) and inicio + 1 < len(blocos):
        notas.append("fim da seção não identificado com certeza — considerado "
                     "até o fim do documento; conferir na planilha")
    return blocos[inicio + 1:fim]


# ---------------------------------------------------------------------------
# Extração das NCs dentro da seção
# ---------------------------------------------------------------------------

def separar_campos(texto):
    """Divide o texto de uma NC em (descrição, norma, solução) pelos rótulos.
    Sem rótulos: descrição = texto completo e a norma é buscada por regex."""
    rotulos = [(r, "norma") for r in ROTULOS_NORMA] + \
              [(r, "solucao") for r in ROTULOS_SOLUCAO]
    # localiza cada rótulo (na forma normalizada) preservando o texto original
    norm = normalizar(texto)
    posicoes = []  # (posição na string normalizada, tamanho, campo)
    for rotulo, campo in rotulos:
        for m in re.finditer(r"(?:^|[.;]\s*)(" + re.escape(rotulo) + r")\s*[:\-–]",
                             norm):
            ini = m.start(1)
            posicoes.append((ini, m.end() - ini, campo))
    # normalizar() colapsa espaços, então os índices normalizado/original
    # coincidem apenas se o texto original também estiver colapsado:
    original = re.sub(r"\s+", " ", texto).strip()
    if len(normalizar(original)) != len(original):
        # acentos removidos não mudam o comprimento; diferenças indicam
        # caracteres especiais — recua para o caminho sem rótulos
        posicoes = []
    campos = {"norma": "", "solucao": ""}
    if posicoes:
        posicoes.sort()
        descricao = original[:posicoes[0][0]].strip(" ;:-–")
        for k, (ini, desloc, campo) in enumerate(posicoes):
            fim = posicoes[k + 1][0] if k + 1 < len(posicoes) else len(original)
            valor = original[ini + desloc:fim].strip(" ;:-–")
            if valor and not campos[campo]:
                campos[campo] = valor
    else:
        descricao = original
    if not campos["norma"]:
        achadas = [a.strip() for a in RE_NORMA.findall(descricao or original) if a.strip()]
        if achadas:
            vistos, unicas = set(), []
            for a in achadas:
                chave = normalizar(a)
                if chave not in vistos:
                    vistos.add(chave)
                    unicas.append(a)
            campos["norma"] = "; ".join(unicas)
    return (descricao or original,
            campos["norma"] or NAO_CONSTA,
            campos["solucao"] or NAO_CONSTA)


def inicio_item(texto):
    """(nível 0/1, texto sem marcador) se a linha inicia um item; senão None."""
    m = RE_ITEM_NUMERADO.match(texto)
    if m:
        return m.group(1).count(".") + 1, m.group(2)
    primeiro = texto[:1]
    if primeiro in MARCADORES_NIVEL1:
        return -1, texto[1:].strip()      # -1 = "um nível abaixo do anterior"
    if primeiro in MARCADORES_NIVEL0:
        return 0, texto[1:].strip()
    m = RE_LETRA.match(texto)
    if m:
        return -1, m.group(1)
    return None


CABECALHOS_TABELA = {
    "descricao": ("nao conformidade", "descricao", "descricao da nao conformidade",
                  "nc", "irregularidade"),
    "norma": ("norma", "item da norma", "embasamento", "referencia normativa",
              "base normativa", "item da norma tecnica"),
    "solucao": ("solucao", "solucao proposta", "acao corretiva", "recomendacao",
                "medida corretiva", "acao proposta"),
}


def parsear_tabela(tabela, arquivo, notas):
    """Tabela com cabeçalho reconhecível vira NCs estruturadas; senão, None."""
    if not tabela or len(tabela) < 2:
        return None
    cabecalho = [normalizar(c) for c in tabela[0]]
    colunas = {}
    for campo, apelidos in CABECALHOS_TABELA.items():
        for idx, celula in enumerate(cabecalho):
            if any(a == celula or a in celula for a in apelidos):
                colunas[campo] = idx
                break
    if "descricao" not in colunas:
        return None
    ncs = []
    for linha in tabela[1:]:
        def celula(campo):
            idx = colunas.get(campo)
            return linha[idx].strip() if idx is not None and idx < len(linha) else ""
        descricao = celula("descricao")
        if not descricao or RE_FOTO.match(descricao):
            continue
        revisar = texto_ilegivel(descricao)
        if revisar:
            notas.append(f"trecho ilegível em tabela: {descricao[:60]!r}")
        ncs.append(NC(descricao=descricao if not revisar else f"{REVISAR} {descricao}",
                      norma=celula("norma") or NAO_CONSTA,
                      solucao=celula("solucao") or NAO_CONSTA,
                      nivel=0, arquivo=arquivo, revisar=revisar))
    return ncs or None


def parsear_secao(blocos_secao, arquivo, notas):
    """Extrai as NCs (com hierarquia) dos blocos internos da seção."""
    ncs = []
    itens = []          # lista de (nível_bruto, [linhas de texto])
    atual = None
    for bloco in blocos_secao:
        if bloco.tabela is not None:
            estruturadas = parsear_tabela(bloco.tabela, arquivo, notas)
            if estruturadas:
                ncs.extend(estruturadas)
            else:                       # tabela sem cabeçalho: achata em linhas
                for linha in bloco.tabela:
                    for cel in linha:
                        for sub in cel.splitlines():
                            sub = sub.strip()
                            if sub and not RE_FOTO.match(sub):
                                ini = inicio_item(sub)
                                if ini:
                                    atual = [ini[0], [ini[1]]]
                                    itens.append(atual)
                                elif atual:
                                    atual[1].append(sub)
            continue
        texto = bloco.texto
        if RE_FOTO.match(texto):
            continue
        ini = inicio_item(texto)
        if ini:
            atual = [ini[0], [ini[1]]]
            itens.append(atual)
        elif atual is not None:
            atual[1].append(texto)
        # texto antes do primeiro item (preâmbulo da seção) é ignorado

    if not itens and not ncs:
        # Sem marcadores reconhecíveis: cada parágrafo vira uma NC e o log avisa
        notas.append("seção sem itens numerados/marcadores — cada parágrafo foi "
                     "tratado como uma NC; conferir hierarquia manualmente")
        itens = [[0, [b.texto]] for b in blocos_secao
                 if b.tabela is None and not RE_FOTO.match(b.texto)]

    # níveis brutos -> 0 (principal) / 1 (subitem)
    niveis = sorted({n for n, _ in itens if n > 0})
    base = niveis[0] if niveis else 0
    ultimo_principal = None
    for nivel_bruto, linhas in itens:
        texto = " ".join(linhas).strip()
        if not texto:
            continue
        if nivel_bruto == -1:
            nivel = 1 if ultimo_principal is not None else 0
        elif nivel_bruto > base:
            nivel = 1
        else:
            nivel = 0
        descricao, norma, solucao = separar_campos(texto)
        revisar = texto_ilegivel(descricao)
        if revisar:
            notas.append(f"trecho ilegível/ambíguo: {descricao[:60]!r}")
            descricao = f"{REVISAR} {descricao}"
        nc = NC(descricao=descricao, norma=norma, solucao=solucao,
                nivel=nivel, arquivo=arquivo, revisar=revisar)
        if nivel == 1 and ultimo_principal is not None:
            ultimo_principal.subitens.append(nc)
        else:
            ncs.append(nc)
            ultimo_principal = nc
    return ncs


# ---------------------------------------------------------------------------
# Consolidação de duplicatas
# ---------------------------------------------------------------------------

def similares(a, b):
    na, nb = normalizar(a.strip(" .;")), normalizar(b.strip(" .;"))
    if na == nb:
        return True
    # uma descrição contida na outra = mesma NC com redação mais/menos completa
    if len(na) >= 30 and len(nb) >= 30 and (na in nb or nb in na):
        return True
    return difflib.SequenceMatcher(None, na, nb).ratio() >= LIMIAR_DUPLICATA


def consolidar(ncs_por_arquivo):
    grupos = []
    for arquivo, ncs in ncs_por_arquivo:
        for nc in ncs:
            alvo = next((g for g in grupos if similares(g.descricao, nc.descricao)),
                        None)
            if alvo is None:
                alvo = Grupo(descricao=nc.descricao, norma=nc.norma,
                             solucao=nc.solucao, arquivos=[nc.arquivo],
                             revisar=nc.revisar)
                grupos.append(alvo)
            else:
                alvo.absorver(nc)
            for sub in nc.subitens:
                sg = next((g for g in alvo.subgrupos
                           if similares(g.descricao, sub.descricao)), None)
                if sg is None:
                    alvo.subgrupos.append(Grupo(
                        descricao=sub.descricao, norma=sub.norma,
                        solucao=sub.solucao, arquivos=[sub.arquivo],
                        revisar=sub.revisar))
                else:
                    sg.absorver(sub)
    return grupos


# ---------------------------------------------------------------------------
# Saídas: planilha, log, amostra
# ---------------------------------------------------------------------------

CABECALHO_PLANILHA = ["ID", "Não conformidade / Subitem",
                      "Descrição da não conformidade",
                      "Item da norma técnica (NR-10 / ABNT)",
                      "Solução proposta", "Frequência", "Relatórios de origem"]


def linhas_planilha(grupos):
    linhas = []
    for n, g in enumerate(grupos, 1):
        linhas.append([len(linhas) + 1, f"NC {n}", g.descricao, g.norma,
                       g.solucao, len(g.arquivos), "; ".join(g.arquivos)])
        for m, sg in enumerate(g.subgrupos, 1):
            linhas.append([len(linhas) + 1, f"NC {n}.{m} (subitem)",
                           sg.descricao, sg.norma, sg.solucao,
                           len(sg.arquivos), "; ".join(sg.arquivos)])
    return linhas


def gravar_planilha(linhas, caminho):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Não conformidades"
    ws.append(CABECALHO_PLANILHA)
    fundo = PatternFill("solid", fgColor="1F4E79")
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = fundo
        celula.alignment = Alignment(vertical="center", wrap_text=True)
    for linha in linhas:
        ws.append(linha)
    larguras = [6, 20, 70, 32, 60, 12, 42]
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura
    quebra = Alignment(vertical="top", wrap_text=True)
    for linha in ws.iter_rows(min_row=2):
        for celula in linha:
            celula.alignment = quebra
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    ws.freeze_panes = "A2"
    wb.save(caminho)


def gravar_log(caminho, processados, pulados, ocr, notas_por_arquivo,
               total_linhas):
    with open(caminho, "w", encoding="utf-8") as log:
        log.write("LOG DE PROCESSAMENTO — consolidação de não conformidades\n")
        log.write("Categoria: Instalações elétricas em más condições\n")
        log.write("=" * 70 + "\n\n")
        log.write(f"Relatórios processados ({len(processados)}):\n")
        for nome, qtd in processados:
            log.write(f"  - {nome}: {qtd} não conformidade(s) extraída(s)\n")
        log.write(f"\nRelatórios pulados ({len(pulados)}):\n")
        for nome, motivo in pulados:
            log.write(f"  - {nome}: {motivo}\n")
        if ocr:
            log.write(f"\nPDFs que requerem OCR — NÃO extraídos ({len(ocr)}):\n")
            for nome in ocr:
                log.write(f"  - {nome}\n")
        avisos = [(n, a) for n, avs in notas_por_arquivo for a in avs]
        if avisos:
            log.write(f"\nAvisos / trechos a revisar ({len(avisos)}):\n")
            for nome, aviso in avisos:
                log.write(f"  - [{nome}] {aviso}\n")
        extraidas = sum(q for _, q in processados)
        log.write("\n" + "=" * 70 + "\n")
        log.write(f"Total de não conformidades extraídas: {extraidas}\n")
        log.write(f"Linhas na planilha após consolidação: {total_linhas}\n")


def mostrar_amostra(linhas):
    print("\nAMOSTRA — 5 primeiras linhas da planilha:")
    print("-" * 70)
    for linha in linhas[:5]:
        id_, hier, desc, norma, sol, freq, origem = linha
        print(f"[{id_}] {hier}  (frequência: {freq})")
        print(f"    Descrição: {desc[:200]}{'…' if len(desc) > 200 else ''}")
        print(f"    Norma:     {norma[:120]}")
        print(f"    Solução:   {sol[:160]}{'…' if len(sol) > 160 else ''}")
        print(f"    Origem:    {origem}")
        print("-" * 70)


# ---------------------------------------------------------------------------
# Programa principal
# ---------------------------------------------------------------------------

def confirmar(pergunta, automatico):
    if automatico:
        print(f"{pergunta} (s/n): s  [--sim]")
        return True
    try:
        return input(f"{pergunta} (s/n): ").strip().lower() in ("s", "sim", "y")
    except EOFError:
        return False


def main():
    global LIMIAR_DUPLICATA
    parser = argparse.ArgumentParser(
        description="Consolida NCs da categoria 'Instalações elétricas em más "
                    "condições' dos relatórios .pdf/.docx de uma pasta.")
    parser.add_argument("--pasta", default=PASTA_PADRAO,
                        help=f"pasta dos relatórios (padrão: {PASTA_PADRAO})")
    parser.add_argument("--sim", action="store_true",
                        help="responde 'sim' a todas as confirmações (uso em lote)")
    parser.add_argument("--limiar", type=float, default=LIMIAR_DUPLICATA,
                        help="similaridade mínima (0–1) para consolidar duas NCs "
                             f"como duplicatas (padrão: {LIMIAR_DUPLICATA})")
    args = parser.parse_args()
    LIMIAR_DUPLICATA = args.limiar

    if not os.path.isdir(args.pasta):
        sys.exit(f"ERRO: pasta não encontrada: {args.pasta}")

    arquivos = sorted(n for n in os.listdir(args.pasta)
                      if os.path.splitext(n)[1].lower() in (".pdf", ".docx")
                      and not n.startswith("~$"))
    if not arquivos:
        sys.exit(f"Nenhum .pdf ou .docx encontrado em {args.pasta}.")

    print(f"Arquivos encontrados em {args.pasta} ({len(arquivos)}):")
    for n in arquivos:
        print(f"  - {n}")
    if not confirmar("\nEstes são os relatórios corretos? Prosseguir",
                     args.sim):
        sys.exit("Cancelado pelo usuário.")

    processados, pulados, requer_ocr = [], [], []
    notas_por_arquivo, ncs_por_arquivo = [], []

    for nome in arquivos:
        caminho = os.path.join(args.pasta, nome)
        notas = []
        try:
            if nome.lower().endswith(".docx"):
                blocos = blocos_docx(caminho)
            else:
                blocos, total_chars = blocos_pdf(caminho)
                if total_chars < 100:      # sem texto selecionável => escaneado
                    requer_ocr.append(nome)
                    print(f"  [OCR]     {nome} — sem texto selecionável")
                    continue
        except Exception as erro:
            pulados.append((nome, f"falha na leitura: {erro}"))
            print(f"  [ERRO]    {nome} — {erro}")
            continue

        secao = localizar_secao(blocos, notas)
        if secao is None:
            pulados.append((nome, "seção da categoria não encontrada"))
            print(f"  [PULADO]  {nome} — seção não encontrada")
            continue

        ncs = parsear_secao(secao, nome, notas)
        if not ncs:
            pulados.append((nome, "seção encontrada, mas sem NCs reconhecíveis"))
            print(f"  [PULADO]  {nome} — seção sem NCs reconhecíveis")
            continue

        total = len(ncs) + sum(len(nc.subitens) for nc in ncs)
        processados.append((nome, total))
        ncs_por_arquivo.append((nome, ncs))
        if notas:
            notas_por_arquivo.append((nome, notas))
        print(f"  [OK]      {nome} — {total} NC(s) extraída(s)")

    if requer_ocr:
        print(f"\nATENÇÃO: {len(requer_ocr)} PDF(s) escaneado(s) (sem texto "
              "selecionável) — o conteúdo NÃO foi adivinhado:")
        for nome in requer_ocr:
            print(f"  - {nome}")
        print("Para incluí-los, aplique OCR antes (ex.: ocrmypdf) e rode de novo.")
        if not confirmar("Continuar a consolidação SEM esses arquivos",
                         args.sim):
            sys.exit("Cancelado pelo usuário. Nenhum arquivo foi gravado.")

    if not ncs_por_arquivo:
        sys.exit("\nNenhuma não conformidade extraída — nada a consolidar. "
                 "Nenhum arquivo foi gravado.")

    grupos = consolidar(ncs_por_arquivo)
    linhas = linhas_planilha(grupos)
    mostrar_amostra(linhas)
    print(f"Total: {len(linhas)} linha(s) consolidada(s) a partir de "
          f"{sum(q for _, q in processados)} NC(s) em {len(processados)} "
          "relatório(s).")

    if not confirmar("\nO formato está correto? Gravar planilha e log",
                     args.sim):
        sys.exit("Cancelado pelo usuário. Nenhum arquivo foi gravado.")

    caminho_planilha = os.path.join(args.pasta, NOME_PLANILHA)
    caminho_log = os.path.join(args.pasta, NOME_LOG)
    for caminho in (caminho_planilha, caminho_log):
        if os.path.exists(caminho):
            if not confirmar(f"'{os.path.basename(caminho)}' já existe. "
                             "Sobrescrever", args.sim):
                sys.exit("Cancelado para não sobrescrever. Renomeie o arquivo "
                         "existente e rode novamente.")

    gravar_planilha(linhas, caminho_planilha)
    gravar_log(caminho_log, processados, pulados, requer_ocr,
               notas_por_arquivo, len(linhas))
    print(f"\nPlanilha gravada: {caminho_planilha}")
    print(f"Log gravado:      {caminho_log}")


if __name__ == "__main__":
    main()
