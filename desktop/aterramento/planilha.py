"""Etapa 2 — Planilha resumo (.xlsx) a partir do modelo do cliente.

Estratégia: **clonar o modelo** ``Planilha_Medicoes_Padrao.xlsx`` e preencher
apenas o cabeçalho (Local / Data) e uma linha por máquina, preservando toda a
formatação, as fórmulas (RESISTÊNCIA EFETIVA e ADEQUADO?) e a formatação
condicional (verde/vermelho) do modelo.

O valor medido (mΩ) e a resistência do prolongador entram como parâmetros
(``medicoes``); ambos vêm do pacote (.zip) e são confirmáveis no app.

Mapa de colunas do modelo (linha de dados começa em 10):
    B ITEM · C MÁQUINA/EQUIPAMENTO · D SETOR · E VALOR MEDIDO (mΩ)
    F RESISTÊNCIA DO PROLONGADOR (PT) · G RESISTÊNCIA EFETIVA (fórmula)
    H ESTÁ/NÃO ESTÁ ADEQUADO? (fórmula) · I OBSERVAÇÕES
"""

from __future__ import annotations

import io
import os
from copy import copy

import openpyxl
from openpyxl.styles import Border, PatternFill

from .imagens import encaixar
from .modelo import FORA_DE_FAIXA, Pacote
from .recursos import caminho_recurso

MODELO_PADRAO = caminho_recurso("modelos", "Planilha_Medicoes_Padrao.xlsx")

# Célula (mesclada G3:G5) onde fica o logo do cliente no modelo da planilha.
_LOGO_COL, _LOGO_ROW = 6, 2  # G3 em índices 0-based (col=6, row=2)


def _trocar_logo_g3(ws, caminho_logo: str) -> None:
    """Substitui a imagem do logo (na célula mesclada G3) pela do cliente,
    preservando posição e tamanho do modelo."""
    from openpyxl.drawing.image import Image as XLImage

    for i, img in enumerate(list(getattr(ws, "_images", []))):
        ancora = getattr(img, "anchor", None)
        de = getattr(ancora, "_from", None)
        if de is not None and de.col == _LOGO_COL and de.row == _LOGO_ROW:
            try:
                original = img.ref.getvalue() if hasattr(img.ref, "getvalue") else None
                if original is None:
                    return
                nova = XLImage(io.BytesIO(encaixar(caminho_logo, original, "png")))
                nova.anchor = ancora  # reaproveita posição/tamanho do modelo
                ws._images[i] = nova
            except Exception:
                pass
            return

LINHA_LOCAL = 5          # B5 = "Local: ..."  | D5 = "Data medições: ..."
LINHA_DADOS_INICIO = 10  # primeira linha de máquina
LINHA_DADOS_FIM_MODELO = 146  # última linha de exemplo no modelo (137 itens)

PROLONGADOR_PADRAO = 0.2  # mΩ (padrão até definirmos a origem no .zip)


def _num(coluna: int) -> int:
    return coluna


def gerarPlanilhaResumo(
    pacote: Pacote,
    caminho_saida: str,
    *,
    local: str | None = None,
    data: str | None = None,
    medicoes: dict | None = None,
    prolongador_padrao: float = PROLONGADOR_PADRAO,
    logo_cliente: str | None = None,
    modelo: str | None = None,
) -> str:
    """Gera a planilha resumo preenchendo o modelo.

    Parâmetros
    ----------
    pacote        : dados lidos do .zip (:class:`Pacote`).
    caminho_saida : caminho do .xlsx a gerar.
    local         : texto do campo "Local:" (padrão = nome do cliente).
    data          : texto do campo "Data medições:" (padrão = data da inspeção).
    medicoes      : dict ``{id_equipamento: {"valor": float|None,
                    "prolongador": float|None}}`` com os valores medidos.
    prolongador_padrao : valor de prolongador quando não informado por máquina.
    modelo        : caminho de um modelo alternativo (padrão = modelo do cliente).
    """
    modelo = modelo or MODELO_PADRAO
    wb = openpyxl.load_workbook(modelo)
    ws = wb.active

    local = local if local is not None else (pacote.cliente.nome or "")
    if data is None:
        dt = pacote.data_inspecao
        data = dt.strftime("%d/%m/%Y") if dt else ""
    ws.cell(LINHA_LOCAL, 2).value = f"Local: {local}"          # B5
    ws.cell(LINHA_LOCAL, 4).value = f"Data medições: {data}"    # D5

    # Captura o estilo da primeira linha de dados (modelo) para replicar.
    estilos = {}
    for col in range(2, 10):  # B..I
        c = ws.cell(LINHA_DADOS_INICIO, col)
        estilos[col] = (
            copy(c.font), copy(c.fill), copy(c.border), copy(c.alignment),
            c.number_format,
        )

    medicoes = medicoes or {}
    equipamentos = pacote.equipamentos

    for i, e in enumerate(equipamentos):
        r = LINHA_DADOS_INICIO + i
        med = medicoes.get(e.chave, {})
        valor = med.get("valor")
        prolongador = med.get("prolongador", prolongador_padrao)

        # Replica os estilos do modelo na linha.
        for col in range(2, 10):
            cell = ws.cell(r, col)
            font, fill, border, align, numfmt = estilos[col]
            cell.font = copy(font)
            cell.fill = copy(fill)
            cell.border = copy(border)
            cell.alignment = copy(align)
            cell.number_format = numfmt

        # ITEM: número de verdade, exibido com dois dígitos (01, 02, … 10, 11).
        ws.cell(r, 2).value = e.numero or (i + 1)
        ws.cell(r, 2).number_format = "00"
        ws.cell(r, 3).value = e.nome_sem_numero or e.nome    # MÁQUINA
        ws.cell(r, 4).value = e.setor                        # SETOR

        fora_faixa = valor == FORA_DE_FAIXA
        if fora_faixa:
            prolongador = 0                                  # regra do ">2000"

        if valor is not None:
            ws.cell(r, 5).value = valor                      # VALOR MEDIDO
            ws.cell(r, 6).value = prolongador                # PROLONGADOR
            ws.cell(r, 7).value = f'=IF(E{r}=">2000",">2000",(E{r}-F{r}))'
            ws.cell(r, 8).value = (
                f'=IF(OR(G{r}>1000,G{r}=">2000"),"NÃO ESTÁ","ESTÁ")'
            )
            if fora_faixa:
                ws.cell(r, 5).number_format = "General"
        else:
            # Sem valor medido (máquina pendente): deixa E..H em branco.
            for col in (5, 6, 7, 8):
                ws.cell(r, col).value = None

        # Observação padrão pela adequação: "ESTÁ" -> em branco; "NÃO ESTÁ"
        # (resistência efetiva > 1000 mΩ) -> "CORREÇÃO CONFORME FLUXOGRAMA".
        obs = ""
        if fora_faixa:
            obs = "CORREÇÃO CONFORME FLUXOGRAMA"
        elif valor is not None:
            efetiva = valor - (prolongador if prolongador is not None else 0)
            if efetiva > 1000:
                obs = "CORREÇÃO CONFORME FLUXOGRAMA"
        ws.cell(r, 9).value = obs
        ws.cell(r, 9).number_format = "General"

    # Limpa as linhas de exemplo remanescentes do modelo (abaixo dos dados).
    ultima_usada = LINHA_DADOS_INICIO + len(equipamentos) - 1
    for r in range(ultima_usada + 1, LINHA_DADOS_FIM_MODELO + 1):
        for col in range(2, 10):
            cell = ws.cell(r, col)
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill()

    # (b) Logo do cliente na célula mesclada G3.
    if logo_cliente and os.path.exists(logo_cliente):
        _trocar_logo_g3(ws, logo_cliente)

    os.makedirs(os.path.dirname(os.path.abspath(caminho_saida)), exist_ok=True)
    wb.save(caminho_saida)
    return caminho_saida
