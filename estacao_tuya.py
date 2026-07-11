#!/usr/bin/env python3
"""
Coleta as leituras de uma estação meteorológica via API da nuvem Tuya
e anexa os dados, com timestamp, em uma planilha Excel (estacao.xlsx),
com uma grandeza por coluna.

Uso:
    python3 estacao_tuya.py                      # coleta e grava no Excel
    python3 estacao_tuya.py --testar-endpoints   # diagnostica o data center

Dependências:
    pip install tuya-connector-python openpyxl

As credenciais podem ser sobrescritas por variáveis de ambiente
(TUYA_ACCESS_ID, TUYA_ACCESS_SECRET, TUYA_DEVICE_ID, TUYA_ENDPOINT),
o que é recomendado se este arquivo for versionado em repositório público.
"""

import os
import socket
import sys
from datetime import datetime
from pathlib import Path

# A Tuya valida a região do IP de origem e costuma rejeitar IPv6 com o erro
# 1114 ("your ip don't have access to this API"). Forçamos IPv4 em todas as
# conexões. Defina TUYA_PERMITIR_IPV6=1 para desativar este ajuste.
if os.environ.get("TUYA_PERMITIR_IPV6") != "1":
    _getaddrinfo_original = socket.getaddrinfo

    def _getaddrinfo_somente_ipv4(host, port, family=0, *args, **kwargs):
        return _getaddrinfo_original(host, port, socket.AF_INET, *args, **kwargs)

    socket.getaddrinfo = _getaddrinfo_somente_ipv4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tuya_connector import TuyaOpenAPI

# ----------------------------------------------------------------------
# Configuração (Project code: p1752408794336gqwnmw)
# ----------------------------------------------------------------------
ACCESS_ID = os.environ.get("TUYA_ACCESS_ID", "aq7ynn9sve3u9n8gku7w")
ACCESS_SECRET = os.environ.get("TUYA_ACCESS_SECRET", "d22b3801be26467c82301367c1bbd17a")
DEVICE_ID = os.environ.get("TUYA_DEVICE_ID", "eb10ba24e88d3576acvgwm")

# Western America Data Center
ENDPOINT = os.environ.get("TUYA_ENDPOINT", "https://openapi.tuyaus.com")

# Todos os data centers da Tuya, para o modo --testar-endpoints
ENDPOINTS_CONHECIDOS = {
    "Western America": "https://openapi.tuyaus.com",
    "Eastern America": "https://openapi-ueaz.tuyaus.com",
    "Central Europe": "https://openapi.tuyaeu.com",
    "Western Europe": "https://openapi-weaz.tuyaeu.com",
    "India": "https://openapi.tuyain.com",
    "China": "https://openapi.tuyacn.com",
}

# Planilha salva no mesmo diretório do script (funciona igual quando rodada pelo agendador)
XLSX_PATH = Path(__file__).resolve().parent / "estacao.xlsx"
NOME_ABA = "Estação"

TIMESTAMP_COL = "timestamp"

# Dicas para os erros mais comuns da API Tuya, indexadas pelo campo "code"
DICAS_ERRO = {
    1004: "Assinatura inválida: confira o Access Secret e sincronize o relógio do computador.",
    1010: "Token inválido: geralmente o endpoint não corresponde ao data center do projeto.",
    1106: "Sem permissão: verifique se o device está vinculado a este projeto no console Tuya.",
    1109: "Endpoint errado: confira o data center do projeto no console Tuya.",
    1114: "IP de origem rejeitado: verifique a 'Cloud Authorization IP Allowlist' do projeto "
    "no console Tuya; se o IP mostrado for IPv6, force IPv4.",
    2007: "Data center incorreto para estas credenciais: ajuste o ENDPOINT.",
    28841105: "Assinatura do plano IoT Core expirou: renove em Cloud > projeto > Service API.",
}


def _detalhar_erro(resposta: dict, contexto: str) -> str:
    code = resposta.get("code")
    msg = f"{contexto} (code={code}): {resposta.get('msg')}"
    dica = DICAS_ERRO.get(code)
    return f"{msg}\nDica: {dica}" if dica else msg


def coletar_status() -> dict:
    """Conecta na nuvem Tuya e devolve {codigo: valor} com todos os status do device."""
    api = TuyaOpenAPI(ENDPOINT, ACCESS_ID, ACCESS_SECRET)
    token = api.connect()
    if not token.get("success"):
        raise RuntimeError(_detalhar_erro(token, "Falha ao obter token na API Tuya"))

    resposta = api.get(f"/v1.0/iot-03/devices/{DEVICE_ID}/status")
    if not resposta.get("success"):
        raise RuntimeError(_detalhar_erro(resposta, "Falha ao consultar o device"))

    # resposta["result"] é uma lista como:
    # [{"code": "va_temperature", "value": 215}, {"code": "va_humidity", "value": 60}, ...]
    return {item["code"]: item["value"] for item in resposta.get("result", [])}


def _criar_planilha(colunas: list):
    """Cria a pasta de trabalho com o cabeçalho formatado (uma grandeza por coluna)."""
    wb = Workbook()
    ws = wb.active
    ws.title = NOME_ABA

    fonte = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fundo = PatternFill(fill_type="solid", fgColor="1F3864")
    alinhamento = Alignment(horizontal="center", vertical="center")
    borda = Border(bottom=Side(style="thin"))

    for indice, nome in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=indice, value=nome)
        celula.font = fonte
        celula.fill = fundo
        celula.alignment = alinhamento
        celula.border = borda
        letra = celula.column_letter
        ws.column_dimensions[letra].width = 21 if nome == TIMESTAMP_COL else max(len(nome) + 2, 10)

    ws.freeze_panes = "A2"
    return wb, ws


def gravar_xlsx(leituras: dict) -> None:
    """Anexa uma linha na planilha; cria o arquivo com cabeçalho na primeira execução."""
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        ws = wb[NOME_ABA] if NOME_ABA in wb.sheetnames else wb.active
        colunas = [celula.value for celula in ws[1] if celula.value is not None]

        # Se o device passou a reportar códigos novos, avisa mas não desalinha as colunas
        novos = sorted(set(leituras) - set(colunas))
        if novos:
            print(
                f"Aviso: códigos novos ignorados (não estão no cabeçalho): {novos}",
                file=sys.stderr,
            )
    else:
        colunas = [TIMESTAMP_COL] + sorted(leituras)
        wb, ws = _criar_planilha(colunas)

    linha = {TIMESTAMP_COL: agora, **leituras}
    ws.append([linha.get(nome) for nome in colunas])

    try:
        wb.save(XLSX_PATH)
    except PermissionError:
        raise RuntimeError(
            f"Não foi possível salvar {XLSX_PATH}: o arquivo está aberto no Excel. "
            "Feche a planilha e rode o script de novo."
        ) from None

    print(f"{agora} - {len(leituras)} leituras gravadas em {XLSX_PATH}")


def testar_endpoints() -> None:
    """Tenta obter o token em cada data center e mostra qual aceita as credenciais."""
    for nome, url in ENDPOINTS_CONHECIDOS.items():
        api = TuyaOpenAPI(url, ACCESS_ID, ACCESS_SECRET)
        try:
            token = api.connect()
        except Exception as exc:  # noqa: BLE001 - queremos seguir testando os demais
            print(f"[FALHA] {nome} ({url}): erro de conexão: {exc}")
            continue
        if token.get("success"):
            print(f"[OK]    {nome} ({url}): token obtido — use este endpoint!")
        else:
            print(f"[FALHA] {nome} ({url}): code={token.get('code')} - {token.get('msg')}")


def main() -> None:
    if "--testar-endpoints" in sys.argv:
        testar_endpoints()
        return

    leituras = coletar_status()
    if not leituras:
        print("Aviso: o device não retornou nenhum status.", file=sys.stderr)
        return
    gravar_xlsx(leituras)


if __name__ == "__main__":
    main()
